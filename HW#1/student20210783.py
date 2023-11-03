#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook('student.xlsx')
sheet = wb['Sheet1']

stdNum = 74
totals = []

for row in sheet.iter_rows(2, stdNum + 1):
    midterm = row[2].value
    final = row[3].value
    homework = row[4].value
    attendance = row[5].value

    total = midterm * 0.3 + final * 0.35 + homework * 0.34 + attendance
    totals.append(total)

totals.sort(reverse=True)

totalA = int(stdNum * 0.3)
totalB = int(stdNum * 0.4)
countA = totalA
countB = totalB
countC = 0

A_boundary = totals[min(totalA - 1, len(totals) - 1)]
B_boundary = totals[min(totalA + totalB - 1, len(totals) - 1)-1]

for i, total in enumerate(totals):
    if total < 40:
        grade = 'F'
    elif total >= A_boundary:
        if countA > int(totalA*0.5):
            grade = 'A+'
        else:
            grade = 'A'
        countA -= 1
    elif total >= B_boundary:
        if countB > int(totalB*0.5):
            grade = 'B+'
        else:
            grade = 'B'
        countB -= 1
    else:
        countC += 1
        grade = 'C'

    # C+ 처리...
    # if grade == 'C' and countC

    for row in sheet.iter_rows(2, stdNum + 1):
        student_id = row[0].value
        midterm = row[2].value
        final = row[3].value
        homework = row[4].value
        attendance = row[5].value

        student_total = midterm * 0.3 + final * 0.35 + homework * 0.34 + attendance

        if student_total == total:
            sheet.cell(row[0].row, 7, total)
            sheet.cell(row[0].row, 8, grade)


wb.save('student.xlsx')
wb.close()
